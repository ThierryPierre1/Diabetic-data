import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook

### Load Data
Data = pd.read_csv('Data/diabetic_data.csv')

## Transforming the data
Data['change'] = Data['change'].str.replace('Ch', 'change')
# Defining a function to group the values in a column by a specified range
def group_by_range(df, col_name, n):
    df_sorted = df.sort_values(col_name)
    min_val = df_sorted[col_name].min()
    max_val = df_sorted[col_name].max()
    num_groups = int((max_val - min_val) / n) + 1
    group_labels = [f"{min_val + i*n}-{min_val + (i+1)*n-1}" for i in range(num_groups)]
    df_sorted[col_name] = pd.cut(df_sorted[col_name], bins=num_groups, labels=group_labels, include_lowest=True)
    return df_sorted
# Calling the function to group the values in the specified columns by a specified range
Data = group_by_range(Data, 'time_in_hospital', 3)
Data = group_by_range(Data, 'num_lab_procedures', 5)
Data = group_by_range(Data, 'num_medications', 5)
Data = group_by_range(Data, 'number_outpatient', 5)
Data = group_by_range(Data, 'number_emergency', 5)
Data = group_by_range(Data, 'number_inpatient', 3)
Data = group_by_range(Data, 'number_diagnoses', 3)
# Remapping the admission_type_id, discharge_disposition_id, and admission_source_id columns with the data dictionaries in the IDs_mapping.xlsx file
file_path = "data\IDs_mapping.xlsx"
wb = load_workbook(filename = file_path)
DataDictionaries = {}
for sheet_name in wb.sheetnames:
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    dict_name = sheet_name.replace(" ", "_").lower()
    data_dict = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))
    DataDictionaries[dict_name] = data_dict
for dict_name, data_dict in DataDictionaries.items():
    matching_cols = [col for col in Data.columns if col == dict_name]
    for col in matching_cols:
        Data[col] = Data[col].map(data_dict)
# Reformatting the 'diag_1', 'diag_2', and 'diag_3' columns into their ICD-9 code descriptions
# Replacing the 'E' values in the 'diag_1', 'diag_2', and 'diag_3' columns to '10' to make conversion to float possible
Data['diag_1'] = Data['diag_1'].astype(str).str.replace('E', '10')
Data['diag_2'] = Data['diag_2'].astype(str).str.replace('E', '10')
Data['diag_3'] = Data['diag_3'].astype(str).str.replace('E', '10')
# Replacing the 'E' values in the 'diag_1', 'diag_2', and 'diag_3' columns to '10' to make conversion to float possible
Data['diag_1'] = Data['diag_1'].astype(str).str.replace('V', '20')
Data['diag_2'] = Data['diag_2'].astype(str).str.replace('V', '20')
Data['diag_3'] = Data['diag_3'].astype(str).str.replace('V', '20')
# Converting the 'diag_1', 'diag_2', and 'diag_3' columns to float
Data['diag_1'] = Data['diag_1'].apply(lambda x: np.ceil(float(x)) if x != '?' else x)
Data['diag_2'] = Data['diag_2'].apply(lambda x: np.ceil(float(x)) if x != '?' else x)
Data['diag_3'] = Data['diag_3'].apply(lambda x: np.ceil(float(x)) if x != '?' else x)
# Defining a dictionary that maps ICD code ranges to their associated descriptions
icd_code_ranges = {
    range(1, 140): 'INFECTIOUS AND PARASITIC DISEASES',
    range(140, 240): 'NEOPLASMS',
    range(240, 280): 'ENDOCRINE, NUTRITIONAL AND METABOLIC DISEASES, AND IMMUNITY DISORDERS',
    range(280, 290): 'DISEASES OF THE BLOOD AND BLOOD-FORMING ORGANS',
    range(290, 320): 'MENTAL, BEHAVIORAL AND NEURODEVELOPMENTAL DISORDERS',
    range(320, 390): 'DISEASES OF THE NERVOUS SYSTEM AND SENSE ORGANS',
    range(390, 460): 'DISEASES OF THE CIRCULATORY SYSTEM',
    range(460, 520): 'DISEASES OF THE RESPIRATORY SYSTEM',
    range(520, 580): 'DISEASES OF THE DIGESTIVE SYSTEM',
    range(580, 630): 'DISEASES OF THE GENITOURINARY SYSTEM',
    range(630, 680): 'COMPLICATIONS OF PREGNANCY, CHILDBIRTH, AND THE PUERPERIUM',
    range(680, 710): 'DISEASES OF THE SKIN AND SUBCUTANEOUS TISSUE',
    range(710, 740): 'DISEASES OF THE MUSCULOSKELETAL SYSTEM AND CONNECTIVE TISSUE',
    range(740, 760): 'CONGENITAL ANOMALIES',
    range(760, 780): 'CERTAIN CONDITIONS ORIGINATING IN THE PERINATAL PERIOD',
    range(780, 800): 'SYMPTOMS, SIGNS, AND ILL-DEFINED CONDITIONS',
    range(800, 1000): 'INJURY AND POISONING',
    range(10000, 10999): 'SUPPLEMENTARYCLASSIFICATION OF EXTERNAL CAUSES OF INJURY AND POISONING',
    range(2000, 2099): 'SUPPLEMENTARY CLASSIFICATION OF FACTORS INFLUENCING HEALTH STATUS AND CONTACT WITH HEALTH SERVICES',
}
# Defining a function that converts the values within the 'diag_1', 'diag_2', and 'diag_3' columns to their ICD-9 code descriptions
def icd_code_to_description(icd_code):
    if icd_code == '?':  # if value is '?', return it unchanged
        return icd_code
    for code_range, description in icd_code_ranges.items():
        if int(icd_code) in code_range:
            return description
    return 'UNKNOWN'
# Applying the icd_code_to_description function to the 'diag_1', 'diag_2', and 'diag_3' columns
Data['diag_1'] = Data['diag_1'].apply(icd_code_to_description)
Data['diag_2'] = Data['diag_2'].apply(icd_code_to_description)
Data['diag_3'] = Data['diag_3'].apply(icd_code_to_description)

# Alternative method of ICD Code Conversion
"""
### Reformatting the 'diag_1', 'diag_2', 'diag_3', columns into their ICD-9 code descriptions
# Loading both ICD9 Code List and merging them into one dataframe
ICD9Code_1 = pandas.read_csv('ICD-9 Codes\CMS32_DESC_LONG_SHORT_DX.xlsx')
ICD9Code_2 = pandas.read_csv('ICD-9 Codes\CMS32_DESC_LONG_SHORT_SG.xlsx')
ICD9CodeList = pandas.concat([ICD9Code_1, ICD9Code_2], ignore_index=True)
# Removing decimal points within the 'diag_1', 'diag_2', 'diag_3', columns to match the ICD9 Code List
Data['diag_1'] = Data['diag_1'].astype(str).str.replace('.', '')
Data['diag_2'] = Data['diag_2'].astype(str).str.replace('.', '')
Data['diag_3'] = Data['diag_3'].astype(str).str.replace('.', '')
Data['diag_1'] = Data['diag_1'].replace(ICD9CodeList .set_index('DIAGNOSIS CODE')['SHORT DESCRIPTION'])
Data['diag_2'] = Data['diag_2'].replace(ICD9CodeList .set_index('DIAGNOSIS CODE')['SHORT DESCRIPTION'])
Data['diag_3'] = Data['diag_3'].replace(ICD9CodeList .set_index('DIAGNOSIS CODE')['SHORT DESCRIPTION'])
"""

### Creating data dictionaries for specific categorical columns 

# Assigning the desired columns to a list
columns = ['race', 'gender', 'age', 'weight', 'admission_type_id', 'discharge_disposition_id', 
           'admission_source_id', 'time_in_hospital', 'payer_code', 'medical_specialty', 'num_lab_procedures', 'num_procedures', 
           'num_medications', 'number_outpatient', 'number_emergency', 'number_inpatient', 'diag_1', 'diag_2', 'diag_3', 
           'number_diagnoses', 'max_glu_serum', 'A1Cresult', 'metformin', 'repaglinide', 'nateglinide', 'chlorpropamide', 
           'glimepiride', 'acetohexamide', 'glipizide', 'glyburide', 'tolbutamide', 'pioglitazone', 'rosiglitazone', 
           'acarbose', 'miglitol', 'troglitazone', 'tolazamide', 'examide', 'citoglipton', 'insulin', 'glyburide-metformin', 
           'glipizide-metformin', 'glimepiride-pioglitazone', 'metformin-rosiglitazone', 'metformin-pioglitazone', 'change', 
           'diabetesMed', 'readmitted']
# Defining a function to create data dictionaries for the desired columns
def create_data_dictionary(Data, columns):
    data_dictionaries = {}
    for col in columns:
        unique_values = Data[col].unique()
        data_dict = dict(zip(unique_values, range(len(unique_values))))
        data_dictionaries[col] = data_dict
        print(f"Data Dictionary for df.{col}:")
        print(data_dict)
        print()
    return data_dictionaries
# Calling the function to create the data dictionaries
data_dictionaries = create_data_dictionary(Data, columns)

### Cleaning the dataset

# Replacing '?' and empty cells with NaN 
Data.replace('?', pd.np.nan, inplace=True)
Data.replace('', pd.np.nan, inplace=True)
# Counting the number of NaN values remaining in each column to ensure dataset was cleaned properly
print(Data.isna().sum())
# Reformatting Categorical Columns with their corresponding dictionary values
for col, data_dict in data_dictionaries.items():
    Data[col].replace(data_dict, inplace=True)
# Exporting the data dictionaries to an Excel file
writer = pd.ExcelWriter('data/diabetic_data_dictionaries.xlsx')
for col, data_dict in data_dictionaries.items():
    pd.DataFrame.from_dict(data_dict, orient="index").to_excel(writer, sheet_name=col)
writer.save()
# Removing the unwanted columns
Data = Data [['encounter_id', 'patient_nbr', 'race', 'gender', 'age', 'weight', 'admission_type_id', 'discharge_disposition_id', 
           'admission_source_id', 'time_in_hospital', 'payer_code', 'medical_specialty', 'num_lab_procedures', 'num_procedures', 
           'num_medications', 'number_outpatient', 'number_emergency', 'number_inpatient', 'diag_1', 'diag_2', 'diag_3', 
           'number_diagnoses', 'max_glu_serum', 'A1Cresult', 'insulin', 'metformin', 'change', 'diabetesMed', 'readmitted']]

### Exporting the cleaned dataset
Data.to_csv('data/diabetic_data_cleaned.csv', index=False)